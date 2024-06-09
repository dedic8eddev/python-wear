import os
from copy import copy
from datetime import datetime
from tempfile import NamedTemporaryFile
from urllib.parse import urlparse

from marshmallow import ValidationError, fields, pre_load
from openpyxl import Workbook
from psycopg2 import sql
from pyramid.httpexceptions import HTTPInternalServerError

from spynl_schemas import Schema, lookup

from spynl.locale import SpynlTranslationString as _

from spynl.main.exceptions import SpynlException
from spynl.main.serial.file_responses import format_value


class BadRedshiftURI(Exception):
    """The uri should contain a username, password, hostname, port and dbname."""


class RedshiftConnectionError(SpynlException):
    """Could not connect to a database."""

    http_escalate_as = HTTPInternalServerError

    def __init__(self):
        # . it, es, de, fr
        message = _('redshift-connection-error')
        super().__init__(message=message)


def parse_connection_string(uri):
    """Return a dictionary with keyword arguments for psycopg2.connect()"""
    options = urlparse(uri)
    dsn = dict(
        dbname=options.path[1:],
        user=options.username,
        password=options.password,
        port=options.port,
        host=options.hostname,
    )
    if not all(dsn.values()):
        raise BadRedshiftURI('bad connection string')
    return dsn


def debug_query(cursor, query):
    if os.environ.get('DEBUG', False):
        import sqlparse

        print(sqlparse.format(cursor.mogrify(query), reindent=True))


def _build_dependencies(columns, dependency_mapping):
    columns = set(columns)
    for c, dependencies in dependency_mapping.items():
        if c in columns:
            columns.remove(c)
            columns |= dependencies
    return columns


def _build_where(where):
    """
    Format the where clause.

    where is a dictionary where the values can be scalars, lists of scalars.
    It recursively handled dictionaries and lists of dictionaries.

    WHERE "tenant" IN (1, 2)
        AND "brand" = 'nike'
        AND "warehouse" IN (51, 52)
        AND (("season" = 'zomer'AND "_year" = '04')
             OR
             ("season" = 'basis' AND "_year" = '00'))
    """

    def _build(where):
        where_clause = []
        for key, value in where.items():
            if not value and isinstance(value, (list, dict)):
                continue

            # a nested structure. We don't care about the key, we care about
            # the key value pairs inside the dictionary.
            if isinstance(value, dict):
                where_clause.append(_build(value))

            # a list of dicts. We don't care about the key, each dictionary
            # has is AND internally and has an OR relationship with the others.
            elif isinstance(value, list) and isinstance(value[0], dict):
                where_clause.append(
                    sql.SQL('({})').format(
                        sql.SQL(' OR ').join(
                            sql.SQL('({})').format(_build(v)) for v in value
                        )
                    )
                )

            # a list of scalars. We generate an IN clause
            elif isinstance(value, list):
                where_clause.append(
                    sql.SQL('{key} IN ({values})').format(
                        key=sql.Identifier(key),
                        values=sql.SQL(', ').join(sql.Literal(v) for v in value),
                    )
                )

            elif key == 'startDate':
                if isinstance(value, datetime):
                    value = int(value.timestamp())
                where_clause.append(
                    sql.SQL('"timestamp" >= {value}').format(
                        key=sql.Identifier(key), value=sql.Literal(value)
                    )
                )

            elif key == 'endDate':
                if isinstance(value, datetime):
                    value = int(value.timestamp())
                where_clause.append(
                    sql.SQL('"timestamp" <= {value}').format(
                        key=sql.Identifier(key), value=sql.Literal(value)
                    )
                )

            # a normal key = value clause
            else:
                where_clause.append(
                    sql.SQL('{key} = {value}').format(
                        key=sql.Identifier(key), value=sql.Literal(value)
                    )
                )

        if where_clause:
            return sql.SQL(' AND ').join(where_clause)
        return sql.SQL('')

    return sql.SQL(' WHERE ') + _build(where)


def _build_sort(sort, aliases=None):
    """
    Format a sort.

    ORDER BY "tenantname" ASC, "n_sold" DESC
    """
    if aliases is None:
        aliases = {}

    def parse_sort_item(s):
        if isinstance(s, str):
            col, dir = s, 'ASC'
        else:
            col, dir = s[0], s[1]
        return sql.Identifier(col), sql.SQL(dir)

    return sql.SQL(' ORDER BY ') + sql.SQL(', ').join(
        [sql.SQL('{} {}').format(*parse_sort_item(s)) for s in sort]
    )


def _build_select_column(column, aliases):
    if column == 'collection':
        return COLLECTION_SQL
    return sql.SQL('{} as {}').format(
        sql.Identifier(column), sql.Identifier(aliases.get(column, column))
    )


def _build_group_by(columns):
    """Format group by

    GROUP BY "supplier", "tenantname"
    """
    return sql.SQL(' GROUP BY ') + sql.SQL(', ').join(
        sql.Identifier(c) for c in columns
    )


COLLECTION_SQL = sql.SQL(
    "COALESCE(\"season\",'') || '-' || " "COALESCE(\"_year\",'') as \"collection\""
)


def build_filter_values(columns, where, aliases=None):
    if aliases is None:
        aliases = {}

    selects = []
    for c in columns:
        if c == 'collection':
            select = sql.SQL(
                "COALESCE(\"season\",'') " "|| '-' || " "COALESCE(\"_year\",'') "
            )
        else:
            select = sql.Identifier(c)

        selects.append(
            sql.SQL(
                'SELECT DISTINCT {} as key, {} as value from transactions {}'
            ).format(
                sql.Literal(aliases.get(c, c)),
                select,
                _build_where(where) if where else sql.SQL(''),
            )
        )
    query = sql.SQL(' UNION ALL ').join(selects)
    query += sql.SQL(' ORDER BY value')
    return query


class CollectionSchema(Schema):
    season = fields.String(required=True)
    _year = fields.String(required=True)

    @pre_load
    def pre_load(self, data, **kwargs):
        try:
            season, year = data.split('-', 1)
        except ValueError:
            raise ValidationError('Must be a string in form of "summer-2018"')
        return dict(season=season, _year=year)


def prepare_filter_response(data, aliases):
    values = {}
    for row in data:
        key, value = row['key'], row['value']
        column = aliases.get(key, key)
        values.setdefault(column, [])
        if value is not None:
            values[column].append(value)
    return values


def revert_back_to_camelcase(data, aliases):
    # Redshift will lowercase any identifier so if we are running
    # against an actual redshift db instead of postgres we need to get back
    # the camelcase identifiers.
    try:
        if isinstance(data, dict):
            return {aliases[k]: v for k, v in data.items()}
        else:
            return [{aliases[k]: v for k, v in row.items()} for row in data]
    except KeyError:
        # If we get a KeyError it means the column names were not changed
        # by the database or database driver.
        return data


def generate_excel_report(parameters, data, totals):
    """generate an excel report for article status"""
    # NOTE indexing on excel starts from 1.

    groups = parameters['groups']
    fields = parameters['fields_']
    metadata = parameters['columnMetadata']

    wb = Workbook()
    ws = wb.active
    ws.title = parameters['report_name']

    # copy style from the first field, because using Font(bold=True) uses different
    # defaults for font name etc
    bold = copy(ws['A1'].font)
    bold.bold = True

    for i, value in enumerate([*groups, *fields], start=1):
        ws.cell(
            row=1, column=i, value=lookup(metadata, f'{value}.label', value)
        ).font = bold

    # start from 2, 1 is the header row
    for row_idx, row in enumerate(data, start=2):
        for column_idx, key in enumerate([*groups, *fields], start=1):
            value, format_ = format_value(metadata, key, row[key])
            ws.cell(row=row_idx, column=column_idx, value=value).number_format = format_

    # The row after the header and data rows
    totals_row = len(data) + 2

    # in the header row we have a label 'totals' and then the total values. if
    # we have groups we merge those columns together and put the 'totals' label
    # in there. Groups such as brand or collection do not have totals. if not
    # we insert a new column.
    if groups:
        ws.merge_cells(
            start_row=totals_row,
            start_column=1,
            end_row=totals_row,
            end_column=len(groups),
        )
        first_non_group_column = len(groups) + 1
    else:
        ws.insert_cols(1)
        first_non_group_column = 2

    ws.cell(row=totals_row, column=1, value='total').font = bold
    for column, field in enumerate(fields, start=first_non_group_column):
        value, format_ = format_value(metadata, field, totals[field])
        cell = ws.cell(row=totals_row, column=column, value=value)
        cell.font = bold
        cell.number_format = format_

    tmp = NamedTemporaryFile()
    wb.save(tmp.name + '.xlsx')
    tmp.seek(0)

    if os.getenv('DEBUG'):
        wb.save('test.xlsx')

    return tmp


def default_filter_values(result, schema):
    for key, value in schema._declared_fields.items():
        if value.metadata.get('column', True):
            if value.metadata.get('include_filter_values', True):
                result.setdefault(value.data_key or key, [])
            else:
                result.setdefault(value.data_key or key, None)
