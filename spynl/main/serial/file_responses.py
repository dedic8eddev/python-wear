import csv
import datetime
import os
from copy import copy
from io import StringIO
from tempfile import NamedTemporaryFile

import pytz
from marshmallow import fields
from openpyxl import Workbook
from pyramid.response import FileIter

from spynl_schemas import Schema, lookup

METADATA_DESCRIPTION = (
    'A way to describe which labels the pdf should use for the columns, and how to '
    'format the values. Any unapplicable data will be ignored, so there is no need to '
    'restrict this to only the columns involved in the query.\n'
    '\n'
    ' The schema is not generated automatically yet, so here is an example value: \n'
    '```\n'
    '{"supplier": {"label": "leverancier", "type": "text"}, "turnOverVelocityAmount": '
    '{"label": "something", "type": "number", "decimals": 2}}\n'
    '```\n'
    'Columns of type "money", "number", "quantity" and "percentage" are formatted in '
    'a specific way. Decimals are only applied to "number" and "percentage" and if no '
    'number of decimals is specified, we show 2 decimals.'
)


class ColumnMetadata(Schema):
    label = fields.String(
        metadata={
            'description': 'The label that should be used for this name or field. The '
            'label should be in the language of the user.'
        }
    )
    type = fields.String(
        metadata={
            'description': 'The type of the column. There is special treatment for '
            "'money', 'number', 'quantity', 'percentage' and 'datetime'. Any other "
            'value will be treated as text.'
        }
    )
    decimals = fields.Integer(
        load_default=2,
        metadata={
            'description': 'The amount of decimals to be displayed for a column. This '
            "is only applied to columns of type 'number' or 'percentage'."
        },
    )


def export_header(data, reference):
    """Return the column names ordered as they are found in the reference."""
    if not data:
        return []
    return sorted(data[0], key=lambda i: reference.index(i) if i in reference else 0)


def export_data(data, header):
    """
    Return the data as a 2 dimensional list.

    The inner lists are ordered according the header.
    """
    return [[row[k] for k in header] for row in data]


def export_excel(header, data, metadata={}, request=None):
    """Export the data as an excel attachment."""
    tmp = NamedTemporaryFile()
    wb = Workbook()
    ws = wb.active

    # copy style from the first field, because using Font(bold=True) uses different
    # defaults for font name etc
    bold = copy(ws['A1'].font)
    bold.bold = True
    for i, value in enumerate(header, start=1):
        ws.cell(
            row=1, column=i, value=lookup(metadata, f'{value}.label', value)
        ).font = bold

    # start from 2, 1 is the header row
    for row_idx, row in enumerate(data, start=2):
        for column_idx, key in enumerate(header, start=1):
            value, format_ = format_value(
                metadata, key, row.get(key, ''), request=request
            )
            ws.cell(row=row_idx, column=column_idx, value=value).number_format = format_

    wb.save(tmp.name + '.xlsx')
    tmp.seek(0)

    if os.getenv('DEBUG'):
        wb.save('test.xlsx')

    return tmp


def format_value(metadata, key, value, request=None):
    cell_formats = {
        'money': '€ #,##0.00',
        'quantity': '0',
        'datetime': 'dd-mm-yy h:mm',
        'number': '0{}{}',
        'percentage': '0{}{}%',
    }
    # Excel does not allow tz aware datetimes, so we need to make them naive:
    if isinstance(value, datetime.datetime):
        if request:
            tz = request.cached_user.get('tz', 'Europe/Amsterdam')
        else:
            tz = 'Europe/Amsterdam'
        value = value.astimezone(pytz.timezone(tz))
        value = value.replace(tzinfo=None)
        metadata.setdefault(key, {'type': 'datetime'})
    try:
        column_metadata = metadata[key]
        decimals = column_metadata.get('decimals', 2)
        if (
            value
            and column_metadata['type'] == 'percentage'
            and not isinstance(value, str)
        ):
            # Excel wants percentages to be fractions:
            value = value / 100
        return (
            value,
            cell_formats[column_metadata['type']].format(
                '.' if decimals else '', '0' * decimals
            ),
        )
    except KeyError:
        return value, ''


def serve_excel_response(response, file, filename):
    response.content_disposition = 'attachment; filename=%s' % filename
    response.content_type = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.app_iter = FileIter(file)
    return response


def export_csv(header, data):
    """Export the data as csv as a string."""
    with StringIO() as tmp:
        writer = csv.DictWriter(tmp, fieldnames=header)
        writer.writeheader()
        writer.writerows(data)
        data = tmp.getvalue()

    return data


def serve_csv_response(response, data):
    response.content_type = 'text/csv'
    response.text = data
    return response


def make_pdf_file_response(request, file, filename=None):
    """rewind file and set correct headers"""
    # rewind filepointer to start of the file
    file.seek(0)
    response = request.response
    response.content_type = 'application/pdf'
    if filename:
        response.content_disposition = 'filename="{}.pdf"'.format(filename)
    response.app_iter = FileIter(file)
    return response
